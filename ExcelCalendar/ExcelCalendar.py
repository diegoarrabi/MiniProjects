import sys
from os import path

from subprocess import run

import pandas as pd

from datetime import date
from datetime import timedelta

from config import *

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

# ###################################################################


def clearScreen() -> None:
    print("\x1b[H\x1b[J")
# ######################################


def makeWorkbook(file_name: str, file_path: str) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = file_name
    workbook.save(file_path)
    return workbook
# ######################################


def openWorkbook(file_path: str) -> None:
    run(['open', file_path])
# ######################################


#  ▗▄▄▖▗▄▄▄▖▗▄▖ ▗▄▄▖▗▄▄▄▖
# ▐▌     █ ▐▌ ▐▌▐▌ ▐▌ █
#  ▝▀▚▖  █ ▐▛▀▜▌▐▛▀▚▖ █
# ▗▄▄▞▘  █ ▐▌ ▐▌▐▌ ▐▌ █
# ###################################################################

clearScreen()

# first_str = input("First Month [ex. 'Mar' or 'March']\n: ")
first_str = "Jul"
month_first = monthInteger(first_str)

# last_str = input("\nLast Month [ex. 'Mar' or 'March']\n: ")
last_str = "Aug"
month_last = monthInteger(last_str)


month_firstday = date(YEAR, month_first["index"], 1)
if month_last['index'] < month_first['index']:
    YEAR = YEAR+1
month_lastday = date(YEAR, month_last["index"], month_last["days"])

firstweek_startdate = getWeekStart(month_firstday)
lastweek_enddate = getWeekEnd(month_lastday)

days_inbetween = (lastweek_enddate - firstweek_startdate)
weeks_in_month = round((days_inbetween.days + 1)/7)


print(firstweek_startdate)
print(lastweek_enddate)
print((days_inbetween.days + 1)/7)
print(round((days_inbetween.days + 1)/7))


def create_calendar_dataframe(start_date, weeks_count):
    """Create a calendar DataFrame with proper structure"""

    # Create column headers (days of week)
    weekdays = [weekday_str[7]] + [weekday_str[i+1] for i in range(6)]  # Monday to Sunday

    # Initialize lists to store calendar data
    calendar_data = []

    for week in range(weeks_count):
        week_dates = []
        week_data = []

        for day in range(7):
            day_date = start_date + timedelta(days=(week * 7 + day))
            week_dates.append(day_date.strftime('%Y-%m-%d'))
            week_data.append(day_date.day)

        calendar_data.append(week_data)

    # Create DataFrame
    calendar_df = pd.DataFrame(calendar_data, columns=weekdays)

    return calendar_df


# Generate the calendar DataFrame
calendar_df = create_calendar_dataframe(firstweek_startdate, weeks_in_month)

print("Calendar DataFrame:")
print(calendar_df)

# Create file path for export
# sheet_name, file_path = makeFileName(first_str, last_str)

# Export to Excel with styling

# Export the calendar
# export_styled_calendar(calendar_df, file_path, sheet_name)

# print(f"\nCalendar exported to: {file_path}")

# Optional: Open the file
# openWorkbook(file_path)
